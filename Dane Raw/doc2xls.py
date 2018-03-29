from __future__ import division
import openpyxl
import textract

if __name__=="__main__":
    book = openpyxl.load_workbook("data.xlsx")
    sheet = book.create_sheet("Dane{}".format(len(book.worksheets)))
    text = textract.process("Dane od 11.01.2018 15.02.2018.doc").decode().replace("\n", "")
    splittedtxt = text.split('|')[36:]
    lostrows = 0
    for i in range(0, len(splittedtxt)):
        if splittedtxt[i-i%11] == "  ":
            if not i%11:
                lostrows += 1
        elif i%11 < 10:
            sheet.cell(row=i//11+2-lostrows, column=i%11+1).value = splittedtxt[i]
    book.save("data.xlsx")
